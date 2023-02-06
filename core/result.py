
import openpyxl

from core.utils import excel


class ResultToExcel:

    def __init__(self, excelSavePath, title):

        self.excelSavePath = excelSavePath  # excel的保存路径
        self.excel = excel  # openpyxl.Workbook()的实例话
        self.sheet = self.excel.create_sheet(title=title)  # 创建工作区
        self.Sheet_line = 1  # 表格的行数

    def saveDomain(self,domainsfile):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '域名'
            self.Sheet_line += 1
        with open(domainsfile, 'r') as f:
            domains = f.readlines()
            for domain in domains:
                self.sheet.cell(self.Sheet_line, 1).value = domain
                self.Sheet_line += 1

        self.excel.save(self.excelSavePath)

    def savePort(self, portfile):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = 'ip端口'
            self.Sheet_line += 1
        with open(portfile, 'r') as f:
            ports = f.readlines()
            for port in ports:
                self.sheet.cell(self.Sheet_line, 1).value = port
                self.Sheet_line += 1

        self.excel.save(self.excelSavePath)

    def saveIp(self, ipfile):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = 'ip'
            self.Sheet_line += 1
        with open(ipfile, 'r') as f:
            ips = f.readlines()
            for ip in ips:
                self.sheet.cell(self.Sheet_line, 1).value = ip
                self.Sheet_line += 1

        self.excel.save(self.excelSavePath)

    def saveHttpxTitle(self, titlefile):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = 'title'
            self.Sheet_line += 1
        with open(titlefile, 'r') as f:
            titles = f.readlines()
            for title in titles:
                self.sheet.cell(self.Sheet_line, 1).value = title
                self.Sheet_line += 1

        self.excel.save(self.excelSavePath)
