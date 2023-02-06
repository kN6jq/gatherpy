#导入openpyxl package
import openpyxl

#创建一个工作簿
f = openpyxl.Workbook()
# table = f.active
#or
#创建sheet
# table = f.create_sheet('AD')    	#创建并插入末尾
table = f.active
#可以通过RGB色值设置sheet标签颜色
table.sheet_properties.tabColor = 'FF00FF00'

# #复制一个工作表
# target = f.copy_worksheet(table)
table['A1'] = 'hello'
table.cell(2,1).value = 'world'
#保存文件
f.template = True    #存为模板
f.save('demo.xlsx')